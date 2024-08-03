#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author: leeyoshinari
import os
import sys
import openpyxl

if hasattr(sys, 'frozen'):
    current_path = os.path.dirname(sys.executable)
else:
    current_path = os.path.dirname(os.path.abspath(__file__))


file_names = [n for n in os.listdir(current_path) if n.endswith('.xlsx')]
for file_name in file_names:
    wb = openpyxl.load_workbook(os.path.join(current_path, file_name))
    ws = wb[wb.sheetnames[0]]
    res = []
    for row in ws.rows:
        if row[0].value and row[1].value:
            if'；' in str(row[1].value):
                res.append({"a": str(row[0].value).strip(), 'b': str(row[1].value).strip().split('；')})
            else:
                res.append({"a": str(row[0].value).strip(), 'b': str(row[1].value).strip().split(';')})

    ind = 1
    for r in res:
        a = r['a']
        for rr in r['b']:
            if rr:
                ws.cell(row=ind, column=4, value=a)
                ws.cell(row=ind, column=5, value=rr)
                ind += 1
    wb.save(os.path.join(current_path, file_name))
