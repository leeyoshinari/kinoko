import os
import sys
import time
import copy
import heapq
import sqlite3
import logging.handlers
import traceback
from datetime import datetime, timedelta
import xlrd
from openpyxl import Workbook


if hasattr(sys, 'frozen'):
    current_path = os.path.dirname(sys.executable)
else:
    current_path = os.path.dirname(os.path.abspath(__file__))
logger = logging.getLogger()
formatter = logging.Formatter('%(asctime)s - %(levelname)s - line:%(lineno)d - %(message)s')
logger.setLevel(level=logging.INFO)
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)


detail_excel = os.path.join(current_path, '明细.xlsx')
origin_excel = os.path.join(current_path, '说明.xlsx')
database_path = os.path.join(current_path, "transforAmount.db")
config_path = os.path.join(current_path, 'config.txt')
conn = sqlite3.connect(database_path)
cursor = conn.cursor()
if os.path.exists(config_path):
    with open(config_path, 'r') as f:
        lines = f.readlines()
    for line in lines:
        if 'delta' in line:
            max_delta = int(line.split('=')[1].strip())
else:
    max_delta = 200

cursor.execute('''
    CREATE TABLE IF NOT EXISTS detail (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        isAmount INTEGER NOT NULL DEFAULT 0,
        order_no CHAR(32) NOT NULL,
        xiangci INTEGER NOT NULL,
        ticket_date CHAR(16),
        sub_company CHAR(32),
        sub_company_code CHAR(32),
        transfor_from CHAR(64) NOT NULL,
        product CHAR(32) NOT NULL,
        material_code CHAR(64),
        material_desc TEXT,
        price FLOAT NOT NULL,
        stock INTEGER NOT NULL,
        transfor_sub_company CHAR(32),
        transfor_company CHAR(64),
        transfor_company_code CHAR(32),
        transfor_amount CHAR(16)
    )''')


def read_detail():
    if not os.path.exists(detail_excel):
        raise Exception(f"订单明细文件不存在，文件路径：{detail_excel}")
    if not os.path.exists(origin_excel):
        raise Exception(f"文件不存在，文件路径：{origin_excel}")
    excel = xlrd.open_workbook(detail_excel)
    sheets = excel.sheet_names()
    table = excel.sheet_by_name(sheets[0])
    try:
        for i in range(1, table.nrows):
            if not table.cell_value(i, 32) and not table.cell_value(i, 33): continue
            order_no = table.cell_value(i, 2)
            xiangci = table.cell_value(i, 3)
            ticket_date = table.cell_value(i, 19)
            sub_company = table.cell_value(i, 23)
            sub_company_code = table.cell_value(i, 26)
            transfor_from = table.cell_value(i, 27)
            product = table.cell_value(i, 29)
            material_code = table.cell_value(i, 30)
            material_desc = table.cell_value(i, 31)
            price = table.cell_value(i, 33)
            stock = table.cell_value(i, 32)
            total_amount = table.cell_value(i, 34)
            if round(float(price) * int(stock), 2) != round(float(total_amount), 2):
                logger.error(f"订单明细中的金额计算不正确，订单号：{order_no}，产品组：{product}，物料编码：{material_code}，{price} × {stock} != {total_amount}")
                continue
            if ticket_date:
                if excel.datemode == 0:  # 1900-based
                    date = datetime(1900, 1, 1) + timedelta(days=ticket_date - 2)
                else:
                    date = datetime(1904, 1, 1) + timedelta(days=ticket_date)
                ticket_date = date.strftime('%d/%m/%Y')
            sql = ("INSERT INTO detail (order_no, xiangci, ticket_date, sub_company, sub_company_code, transfor_from, "
                   "product, material_code, material_desc, price, stock) VALUES {};").format((order_no, xiangci, ticket_date,
                    sub_company, sub_company_code, transfor_from, product, material_code, material_desc, float(price), int(stock)))
            cursor.execute(sql)
            conn.commit()
            logger.info(f"正在读取明细：{i} - {order_no} - {xiangci} - {transfor_from} - {product} - {material_code}")
    except:
        logger.error(traceback.format_exc())


def deal_excel():
    if not os.path.exists(origin_excel):
        raise Exception(f"文件不存在，文件路径：{origin_excel}")
    excel = xlrd.open_workbook(origin_excel)
    sheets = excel.sheet_names()
    table = excel.sheet_by_name(sheets[0])
    fail_list = []
    delta = 20
    for i in range(1, table.nrows):
        if not table.cell_value(i, 0) and not table.cell_value(i, 1) and not table.cell_value(i, 2): continue
        company = table.cell_value(i, 0)
        product = table.cell_value(i, 1)
        amount = round(float(table.cell_value(i, 2)) * 10000, 2)
        if amount < 0.1:
            logger.error(f"划拨金额为 0，跳过。{company} - {product}")
            continue
        fail_res = calc_amount((company, product, amount), delta)
        if fail_res:
            fail_list.append(fail_res)

    index = 2
    while len(fail_list) > 0 or delta * index <= max_delta:
        delta1 = delta * index
        index += 1
        if fail_list:
            logger.info(f"正在重试未计算出划拨金额的渠道商......")
        retry_res = []
        while fail_list:
            item = fail_list.pop()
            fail_res = calc_amount(item, delta1)
            if fail_res:
                retry_res.append(fail_res)
        fail_list = copy.deepcopy(retry_res)

    for r in fail_list:
        logger.error(f"划拨金额未计算出，{r[0]} - {r[1]} - {r[2]}")


def calc_amount(res, delta):
    company, product, amount = res
    sql = "select id, price, stock from detail where isAmount = 0 and product = '{}';".format(product)
    cursor.execute(sql)
    result = cursor.fetchall()
    res = buy_items(result, amount, delta)
    fail_list = None
    if res:
        try:
            update_data(res, company)
            buy_price = [price * stock for _, price, stock in res]
            logger.info(f"{company} - {product} - {amount}，计算完成，划拨金额：{round(sum(buy_price), 2)}，误差：{round(amount - sum(buy_price), 2)}")
        except:
            logger.error(traceback.format_exc())
    else:
        fail_list = (company, product, amount)
        logger.warning(f"{company} - {product} - {amount}，未计算出划拨金额，稍后重试~")
    return fail_list


def buy_items(items, amount, tolerance=0):
    items.sort(key=lambda x: -x[1])
    heap = []
    purchased = []
    for ids, price, stock in items:
        max_buy = min(amount // price, stock)
        if max_buy > 0:
            heapq.heappush(heap, (-max_buy * price, max_buy, price, ids))
    while heap:
        _, buy_count, price, ids = heapq.heappop(heap)
        max_affordable = amount // price
        actual_buy = min(buy_count, max_affordable)
        if actual_buy > 0:
            purchased.append((ids, price, actual_buy))
            amount -= actual_buy * price
        if amount <= tolerance:
            return purchased
    return []


def update_data(buy_item, company):
    for ids, price, stock in buy_item:
        sql = "select * from detail where id = {};".format(ids)
        cursor.execute(sql)
        result = cursor.fetchall()
        res = list(result[0])
        if res[12] == stock:
            sql = "update detail set isAmount = 1, transfor_company = '{}', transfor_amount = '{}' where id = {};".format(company, str(round(price * stock, 2)), ids)
            cursor.execute(sql)
            conn.commit()
        if res[12] > stock:
            res[12] = res[12] - stock
            res.pop(0)
            res = res[: -4]
            sql = ("INSERT INTO detail (isAmount, order_no, xiangci, ticket_date, sub_company, sub_company_code, transfor_from, "
                   "product, material_code, material_desc, price, stock) VALUES {};").format(tuple(res))
            cursor.execute(sql)
            conn.commit()
            sql = "update detail set isAmount = 1, stock = {}, transfor_company = '{}', transfor_amount = '{}' where id = {};".format(stock, company, str(round(price * stock, 2)), ids)
            cursor.execute(sql)
            conn.commit()


def write_excel():
    sql = "select * from detail where isAmount = 1;"
    cursor.execute(sql)
    result = cursor.fetchall()
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "分公司"
    sheet["B1"] = "划出渠道商代码"
    sheet["C1"] = "划出渠道商名称"
    sheet["D1"] = "订单号"
    sheet["E1"] = "项次"
    sheet["F1"] = "开票日期"
    sheet["G1"] = "数量"
    sheet["H1"] = "单价"
    sheet["I1"] = "物料编码"
    sheet["J1"] = "物料描述"
    sheet["K1"] = "产品组"
    sheet["L1"] = "划拨后分公司"
    sheet["M1"] = "划入渠道商代码"
    sheet["N1"] = "划入渠道商名称"
    sheet["O1"] = "划拨金额"
    index = 2
    for r in result:
        sheet.cell(row=index, column=1, value=r[5])
        sheet.cell(row=index, column=2, value=r[6])
        sheet.cell(row=index, column=3, value=r[7])
        sheet.cell(row=index, column=4, value=r[2])
        sheet.cell(row=index, column=5, value=r[3])
        sheet.cell(row=index, column=6, value=r[4])
        sheet.cell(row=index, column=7, value=r[12])
        sheet.cell(row=index, column=8, value=r[11])
        sheet.cell(row=index, column=9, value=r[9])
        sheet.cell(row=index, column=10, value=r[10])
        sheet.cell(row=index, column=11, value=r[8])
        sheet.cell(row=index, column=12, value=r[13])
        sheet.cell(row=index, column=13, value=r[15])
        sheet.cell(row=index, column=14, value=r[14])
        sheet.cell(row=index, column=15, value=r[16])
        index += 1

    workbook.save(os.path.join(current_path, time.strftime('%Y%m%d%H%M%S') + '.xlsx'))
    sql = "delete from detail;"
    cursor.execute(sql)
    conn.commit()
    logger.info(f"划分金额 Excel 文件已生成 ~")


if __name__ == '__main__':
    try:
        read_detail()
        deal_excel()
        write_excel()
    except Exception as e:
        logger.error(e)
    time.sleep(1)
    g = input("按回车键继续...")
